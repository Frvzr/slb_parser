__author__ = "Ivan Koposhilov"
__copyright__ = "Copyright (C) 2021 Ivan Koposhilov"
__version__ = "1.0"

import requests
from bs4 import BeautifulSoup
import openpyxl
from datetime import datetime
from urllib3.exceptions import InsecureRequestWarning
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, PatternFill
from settings import URL
import threading
import tkinter as tk

requests.packages.urllib3.disable_warnings(category=InsecureRequestWarning)

big_data = []

def add_user():
    data_entry = f'{gin_entry.get()}: {name_entry.get()}'
    if slb_listbox.curselection():
        lsbx = slb_listbox
        slb_listbox.insert('end', data_entry)
    else:
        lsbx = rs_listbox
        rs_listbox.insert('end', data_entry)
    name_entry.delete(0, 'end')
    gin_entry.delete(0, 'end')
    save_user(lsbx)
    
def del_user():
    if slb_listbox.curselection():
        lsbx = slb_listbox
        select_user = slb_listbox.curselection()
        slb_listbox.delete(select_user)
    else:
        lsbx = rs_listbox
        select_user = rs_listbox.curselection()
        rs_listbox.delete(select_user)
    save_user(lsbx)

def save_user(lsbx):
    if lsbx == slb_listbox:
        with open ('slb_user.txt', 'w+') as f:
            f.writelines("\n".join(lsbx.get(0, 'end')))
    else:
        with open ('rs_user.txt', 'w+') as f:
            f.writelines("\n".join(lsbx.get(0, 'end')))

def show_result(message):
    text.configure(state='normal')
    text.insert('end', message)
    text.configure(state='disabled')

def start_action():
    start_button.config(state=tk.DISABLED)
    thread = threading.Thread(target=load_data)
    thread.start()
    check_thread(thread)
                 
def check_thread(thread):
    if thread.is_alive():
        window.after(100, lambda: check_thread(thread))
    else:
        start_button.config(state=tk.NORMAL)

def load_data():   
    """
    Функция для загрузки в словарь: GIN / имя и в список: необходимые сертификации
    Returns: dict, list
    """
    l = r_var.get()
    if  l == 1:
        lstbx = rs_listbox
    else:
        lstbx = slb_listbox
        
    people_dict = {}
    #people_dict = dict((a, b) for a, b in str(slb_listbox.get(elem).split(': ')) for elem in range(slb_listbox.size()))
    
    for elem in range(lstbx.size()):
        gin_name = str(lstbx.get(elem)).split(': ')
        people_dict[int(gin_name[0])] = gin_name[1]

    connect_quest(people_dict)

def connect_quest(people_dict):
    """
    функция для сбора информаии с Quest

    Args:
        certifications (_list_): список необходимых сертифкаций
        people_dict (_dict_): словарь с табельным номером сотрудника и его ФИО

    Returns:
        _type_: _description_
    """
    
    for gin in people_dict.keys():
        
        name = people_dict[gin]
        
        url = f'{URL}{gin}'

        response = requests.get(url, verify=False)
        soup = BeautifulSoup(response.text, 'html.parser')
        verify_data(soup, name, gin, people_dict)
        
def verify_data(soup, name, gin, people_dict):
    
    try:
        verify_users = soup.find_all('table', class_='altyellow')
    except:
        verify_users = ''
    if verify_users:
        for verify_user in verify_users:  
            try:
                row_user = verify_user.find_all('td')
            except:
                row_user = ''
            try:
                col_user1 = row_user[1].text.strip()
            except:
                col_user1 = ''
            try:
                col_user2 = row_user[5].text.strip()
            except:
                col_user2 = ''
                    
            if col_user1 == name and col_user2 == str(gin):
                tables = soup.find_all('table')[4].contents
                message = f'{name} - данные получены'
            else:
                tables = ''
                message = f'{name} - ошибка данных'
            show_result(message)
    else:
        tables = ''
        message = f'{name} - ошибка данных'
        show_result(message)
        
    collect_data(tables, name, people_dict)
        
        
def collect_data(tables, name, people_dict): 
    global big_data
    data = []
    gin_data = []
    passed_certifications = []

    for item in tables:
        rows = item.find_all('td')
        try:
            col0 = rows[0].text.strip()
        except:
            col0 = ''
        for i in certifications:
            if i == col0:
                col1 = rows[1].text.strip()
                data.append([col0, col1])          
                passed_certifications.append(i)
        
    for element in certifications:
        if element not in passed_certifications:
            data.append([element, 'N/a'])
                
    data = sorted(data,  key=lambda x: x[0])
    for i in data:
        gin_data.append(i[1])
            
    big_data.append([name, *gin_data])
        
    format_file(big_data, certifications, people_dict)

def format_file(big_data, certifications, people_dict):
    date_time = datetime.now()
    date = date_time.strftime('%d-%m-%Y')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Name/Certifications', *certifications])
    
    for row in big_data:
        ws.append(row)
    quantity_users = slb_listbox.size()
    
    tab = Table(displayName="Table1", ref="A1:R{}".format(quantity_users + 1))

    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    for sheet in wb.worksheets:
        for row in sheet:
            for cell in row:
                if cell.value == 'N/a' or cell.value == 'No':
                  cell.fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
                if cell.value in certifications:
                    cell.alignment = Alignment(vertical='center', horizontal='center', text_rotation=90)
    try:                
        ws.column_dimensions['A'].width = 23
        wb.save(f'BOM_{date}.xlsx')
        message = f', данные записаны в файл \n'
    except:
        message = f', ошибка записи в файл \n'
    
    show_result(message)

window = tk.Tk()
window.title('BOM Quest Certifications')
window.geometry('880x600')
window.resizable(0, 0)

gin_label = tk.Label(text='GIN')
gin_label.place(x=10, y=10)

name_label = tk.Label(text='ALIAS')
name_label.place(x=10, y=40)

add_button = tk.Button(text="ADD", height = 1, width=7, command=add_user)
add_button.place(x=130, y=70)

del_button = tk.Button(text='DELETE', command=del_user)
del_button.place(x=200, y=70)

name_entry = tk.Entry(width=33)
name_entry.place(x=50, y=40)

gin_entry = tk.Entry(width=33)
gin_entry.place(x=50, y=10)

slb_listbox = tk.Listbox(height=20, width=40)
slb_listbox.place(x=10, y=125)

rs_listbox = tk.Listbox(height=10, width=40)
rs_listbox.place(x=290, y=125)

cert_listbox = tk.Listbox(height=20, width=40)
cert_listbox.place(x=570, y=125)

slb_label = tk.Label(window, text='SLB USERS')
slb_label.place(x=10, y=100)

rs_label = tk.Label(window, text='RS USERS')
rs_label.place(x=290, y=100)

cert_label = tk.Label(text='CERTIFICATIONS')
cert_label.place(x=650, y=15)

add_cert_button = tk.Button(text="ADD", height = 1, width=7)
add_cert_button.place(x=690, y=70)

del_cert_button = tk.Button(text='DELETE')
del_cert_button.place(x=760, y=70)

cert_entry = tk.Entry(width=40)
cert_entry.place(x=570, y=40)

r_var = tk.IntVar()
r_var.set(0)
slb_rbtn = tk.Radiobutton(text='SLB', variable=r_var, value=0)
slb_rbtn.place(x=600, y=480)
rs_rbtn = tk.Radiobutton(text='RS', variable=r_var, value=1)
rs_rbtn.place(x=600, y=500)

file_dict = {'slb_user.txt': slb_listbox, 'rs_user.txt': rs_listbox, 'certifications.txt': cert_listbox}
for key, values in file_dict.items():
    with open(key) as f:
        lines = f.read().splitlines()
        for i, line in enumerate(lines):
            values.insert(i, line)
            

text = tk.Text(window, state='disable' , height=5, width=65, bg='#E2E8ED')
text.place(x=10, y=480)

start_button = tk.Button(text='START', height=5, width=20, bg='#889AA5', command=start_action)
start_button.place(x=670, y=480)

certifications = [cert_listbox.get(elem) for elem in range(cert_listbox.size())]

if __name__ == "__main__":
    window.mainloop()
    