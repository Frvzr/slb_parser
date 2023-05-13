__author__ = "Ivan Koposhilov"
__copyright__ = "Copyright (C) 2021 Ivan Koposhilov"
__version__ = "2.0"

import requests
from bs4 import BeautifulSoup
import openpyxl
from datetime import datetime
from urllib3.exceptions import InsecureRequestWarning
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, PatternFill
from settings import URL
import threading
import tkinter as tk

requests.packages.urllib3.disable_warnings(category=InsecureRequestWarning)

big_data = []

def save_certification(func):
    def wrapper():
        result = func()
        with open ('data/certifications.txt', 'w+') as f:
            f.writelines("\n".join(cert_listbox.get(0, 'end')))
        return result
    return wrapper

def select_listbox(func):
    def wrapper():
        if rs_listbox.curselection():
            lstbx = rs_listbox
        else:
            lstbx = slb_listbox
        result = func(lstbx)
        return result
    return wrapper

def select_radiobutton(func):
    def wrapper():
        value_rbutton = select_listbox_button.get()
        if  value_rbutton == 1:
            lstbx = rs_listbox
        else:
            lstbx = slb_listbox
        result = func(lstbx)
        return result
    return wrapper

def save_users(func):
    def wrapper():
        result = func()
        if result == slb_listbox:
            file = 'data/slb_user.txt'
        else:
            file = 'data/rs_user.txt'
        with open (file, 'w+') as f:
            f.writelines("\n".join(result.get(0, 'end')))
        return result
    return wrapper  

@save_users
@select_listbox
def add_user(lstbx):
    data_entry = f'{gin_entry.get()}: {name_entry.get()}'
    lstbx.insert('end', data_entry)
    name_entry.delete(0, 'end')
    gin_entry.delete(0, 'end')
    return lstbx

@save_users
@select_listbox 
def del_user(lstbx):
    select_user = lstbx.curselection()
    lstbx.delete(select_user)
    return lstbx
    
@save_certification    
def add_certification():
    certification = f'{cert_entry.get()}' 
    cert_listbox.insert('end', certification)

@save_certification 
def del_certification():
    select_certification = cert_listbox.curselection()
    cert_listbox.delete(select_certification)

@select_radiobutton    
def file_name(lstbx):
    if lstbx == rs_listbox:
        file_name = 'data/rs_user.txt'
    else:
        file_name = 'data/slb_user.txt'
    return file_name

def show_result(message):
    text.configure(state='normal')
    text.insert('end', message)
    text.configure(state='disabled')

def start_action():
    global big_data
    big_data = []
    start_button.config(state=tk.DISABLED)
    thread = threading.Thread(target=load_data)
    text.configure(state='normal')
    text.delete('1.0', 'end')
    text.configure(state='disabled')
    thread.start()
    check_thread(thread)
                 
def check_thread(thread):
    if thread.is_alive():
        window.after(100, lambda: check_thread(thread))
    else:
        start_button.config(state=tk.NORMAL)

@select_radiobutton
def load_data(lstbx):   
    """
    Функция для загрузки в словарь: GIN / имя и в список: необходимые сертификации
    Returns: dict, list
    """
    people_dict = {}
    
    for elem in range(lstbx.size()):
        gin_name = str(lstbx.get(elem)).split(': ')
        people_dict[int(gin_name[0])] = gin_name[1]

    connect_quest(people_dict)

def connect_quest(people_dict):
    """
    функция для сбора информации с Quest

    Args:
        certifications (_list_): список необходимых сертифкаций
        people_dict (_dict_): словарь с табельным номером сотрудника и его ФИО

    Returns:
        _type_: _description_
    """
    
    for gin in people_dict.keys():
        name = people_dict[gin]
        url = f'{URL}{gin}'

        response = requests.get(url, verify=False)
        soup = BeautifulSoup(response.text, 'html.parser')
        verify_data(soup, name, gin)
        
def verify_data(soup, name, gin):
    
    try:
        verify_users = soup.find_all('table', class_='altyellow')
    except:
        verify_users = ''
    if verify_users:
        for verify_user in verify_users:  
            try:
                row_user = verify_user.find_all('td')
            except:
                row_user = ''
            try:
                col_user1 = row_user[1].text.strip()
            except:
                col_user1 = ''
            try:
                col_user2 = row_user[5].text.strip()
            except:
                col_user2 = ''
                    
            if col_user1 == name and col_user2 == str(gin):
                tables = soup.find_all('table')[4].contents
                message = f'{name} - данные получены'
            else:
                tables = ''
                message = f'{name} - ошибка данных'
            show_result(message)
    else:
        tables = ''
        message = f'{name} - ошибка данных'
        show_result(message)
        
    collect_data(tables, name)
   
def collect_data(tables, name): 
    global big_data
    data = []
    gin_data = []
    passed_certifications = []

    for item in tables:
        rows = item.find_all('td')
        try:
            col0 = rows[0].text.strip()
        except:
            col0 = ''
        for i in certifications:
            if i == col0:
                col1 = rows[1].text.strip()
                data.append([col0, col1])          
                passed_certifications.append(i)
        
    for element in certifications:
        if element not in passed_certifications:
            data.append([element, 'N/a'])
                
    data = sorted(data,  key=lambda x: x[0])
    for i in data:
        gin_data.append(i[1])
            
    big_data.append([name, *gin_data])
        
    create_file(certifications)

def get_date():
    date_time = datetime.now()
    date = date_time.strftime('%Y-%m-%d')
    return date

def create_file(certifications):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Name/Certifications', *certifications])
    format_file(ws, wb)
    
@select_radiobutton    
def get_qty_user(lstbx):
    qty_user = lstbx.size()
    return qty_user    
    
def format_file(ws, wb):
    for row in big_data:
        ws.append(row)
    quantity_users = get_qty_user()
    
    tab = Table(displayName="Table1", ref="A1:R{}".format(quantity_users + 1))

    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    ws.column_dimensions['A'].width = 23 
    
    for sheet in wb.worksheets:
        for row in sheet:
            for cell in row:
                if cell.value == 'N/a' or cell.value == 'No':
                  cell.fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
                if cell.value in certifications:
                    cell.alignment = Alignment(vertical='center', horizontal='center', text_rotation=90)
    save_excel(wb)
    
def save_excel(wb):
    segment_name = file_name()
    date = get_date()
    try:
        f_name = f'{segment_name}_{date}.xlsx'       
        wb.save(f_name)
        message = f', данные записаны в файл \n'
    except:
        message = f', ошибка записи в файл \n'
    finally:
        wb.close
        show_result(message)

window = tk.Tk()
window.title('BOM Quest Certifications')
window.geometry('880x600')
window.resizable(0, 0)

gin_label = tk.Label(text='GIN')
gin_label.place(x=10, y=10)

name_label = tk.Label(text='ALIAS')
name_label.place(x=10, y=40)

add_button = tk.Button(text="ADD", height = 1, width=7, command=add_user)
add_button.place(x=130, y=70)

del_button = tk.Button(text='DELETE', command=del_user)
del_button.place(x=200, y=70)

name_entry = tk.Entry(width=33)
name_entry.place(x=50, y=40)

gin_entry = tk.Entry(width=33)
gin_entry.place(x=50, y=10)

slb_listbox = tk.Listbox(height=20, width=40)
slb_listbox.place(x=10, y=125)

rs_listbox = tk.Listbox(height=10, width=40)
rs_listbox.place(x=290, y=125)

cert_listbox = tk.Listbox(height=20, width=40)
cert_listbox.place(x=570, y=125)

slb_label = tk.Label(window, text='SLB USERS')
slb_label.place(x=10, y=100)

rs_label = tk.Label(window, text='RS USERS')
rs_label.place(x=290, y=100)

cert_label = tk.Label(text='CERTIFICATIONS')
cert_label.place(x=650, y=15)

add_cert_button = tk.Button(text="ADD", height = 1, width=7, command=add_certification)
add_cert_button.place(x=690, y=70)

del_cert_button = tk.Button(text='DELETE', command=del_certification)
del_cert_button.place(x=760, y=70)

cert_entry = tk.Entry(width=40)
cert_entry.place(x=570, y=40)

select_listbox_button = tk.IntVar()
select_listbox_button.set(0)
slb_rbtn = tk.Radiobutton(text='SLB', variable=select_listbox_button, value=0)
slb_rbtn.place(x=600, y=480)
rs_rbtn = tk.Radiobutton(text='RS', variable=select_listbox_button, value=1)
rs_rbtn.place(x=600, y=500)

file_dict = {'data/slb_user.txt': slb_listbox, 'data/rs_user.txt': rs_listbox, 'data/certifications.txt': cert_listbox}
for key, values in file_dict.items():
    with open(key) as f:
        lines = f.read().splitlines()
        for i, line in enumerate(lines):
            values.insert(i, line)
            

text = tk.Text(window, state='disable' , height=5, width=65, bg='#E2E8ED')
text.place(x=10, y=480)

start_button = tk.Button(text='START', height=5, width=20, bg='#889AA5', command=start_action)
start_button.place(x=670, y=480)

certifications = [cert_listbox.get(elem) for elem in range(cert_listbox.size())]

if __name__ == "__main__":
    window.mainloop()
    