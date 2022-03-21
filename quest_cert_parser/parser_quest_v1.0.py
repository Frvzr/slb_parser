__author__ = "Ivan Koposhilov"
__copyright__ = "Copyright (C) 2021 Ivan Koposhilov"
__version__ = "1.0"

import requests
from bs4 import BeautifulSoup
import openpyxl
from datetime import datetime
from urllib3.exceptions import InsecureRequestWarning
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, PatternFill
from settings import URL

requests.packages.urllib3.disable_warnings(category=InsecureRequestWarning)


def load_data():   
    """
    Функция для загрузки в словарь: GIN / имя и в список: необходимые сертификации
    Returns: dict, list
    """
    certifications = []
    people_dict = {}

    with open('bom_user_full.txt') as file:
        lines = file.read().splitlines()
    for line in lines:
        key, value = line.split(': ')
        people_dict.update({int(key): value})

    with open('certifications.txt') as file_cert:
        l = file_cert.read().splitlines()
        for line_cert in l:
            certifications.append(line_cert)
    certifications = sorted(certifications)

    return connect_quest(certifications, people_dict)


def connect_quest(certifications, people_dict):
    """
    функция для сбора информаии с Quest

    Args:
        certifications (_list_): список необходимых сертифкаций
        people_dict (_dict_): словарь с табельным номером сотрудника и его ФИО

    Returns:
        _type_: _description_
    """
    big_data = []
    for gin in people_dict.keys():
        
        url = f'{URL}{gin}'

        r = requests.get(url, verify = False)
        soup = BeautifulSoup(r.text, 'html.parser')
        
        try:
            verify_users = soup.find_all('table', class_='altyellow')
        except:
            verify_users = ''
            
        for verify_user in verify_users:  
            row_user = verify_user.find_all('td')
            try:
                col_user1 = row_user[1].text.strip()
            except:
                col_user1 = ''
            try:
                col_user2 = row_user[5].text.strip()
            except:
                col_user2 = ''
        
        if col_user1 == people_dict[gin] and col_user2 == str(gin):
            tables = soup.find_all('table')[4].contents
            print(people_dict[gin], '- данные получены')
        else:
            tables = ''
            print(people_dict[gin], "- ошибка данных")
            
        data = []
        gin_data = []
        passed_certifications = []

        for item in tables:
            rows = item.find_all('td')
            try:
                col0 = rows[0].text.strip()
            except:
                col0 = ''
            for i in certifications:
                if i == col0:
                    col1 = rows[1].text.stri 
                    data.append([col0, col1])          
                    passed_certifications.append(i)   
        
        for element in certifications:
            if element not in passed_certifications:
                data.append([element, 'N/a'])
                
        data = sorted(data,  key=lambda x: x[0])
        for i in data:
            gin_data.append(i[1])
            
        big_data.append([people_dict[gin], *gin_data])
        
    return format_file(big_data, certifications, people_dict)

def format_file(big_data, certifications, people_dict):
    date_time = datetime.now()
    date = date_time.strftime('%d-%m-%Y')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Name/Certifications', *certifications])
    
    for row in big_data:
        ws.append(row)
    quantity_users = len(people_dict)
    
    tab = Table(displayName="Table1", ref="A1:R{}".format(quantity_users + 1))

    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    for sheet in wb.worksheets:
        for row in sheet:
            for cell in row:
                if cell.value == 'N/a' or cell.value == 'No':
                  cell.fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
                if cell.value in certifications:
                    cell.alignment = Alignment(vertical='center', horizontal='center', text_rotation=90)
                    
    ws.column_dimensions['A'].width = 23

    wb.save(f'BOM_{date}.xlsx')

load_data()